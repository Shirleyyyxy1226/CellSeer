"""
readers/cycling/detect.py — Auto-detect cycler format from file content.

Strategy
--------
Read only the column headers, normalize them (lowercase, alphanumeric only),
then score against known cycler signatures using elastic matching:

- Each signature concept is a list of alternatives (OR logic).
  Any alternative matching any header counts as one concept matched.
- Normalization strips spaces, punctuation and units so
  "Chg. Cap.(mAh)", "Charge Capacity(Ah)", "Cap.Chg." all match "chgcap".

Adding a new cycler
-------------------
1. Add an entry to CYCLER_SIGNATURES. Each element is a list of normalized
   substrings — any one of them matching a header counts as a hit.
2. Add a corresponding entry to _CYCLER_READER_MAP.
3. That's it — detect_cycler() picks it up automatically.

Example — adding a hypothetical "Arbin" reader::

    CYCLER_SIGNATURES["arbin"] = [
        ["cycleindex", "cycleid"],
        ["chargecapacity", "chgcap"],
        ["internalresistance", "dcir"],
    ]
    _CYCLER_READER_MAP["arbin"] = ArbinReader
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Type

from cellseer.readers.base import BaseReader


# ---------------------------------------------------------------------------
# Cycler signatures.
# Each entry is a list of *concepts*; each concept is a list of alternative
# normalized substrings (OR within a concept, AND across concepts for score).
#
# Normalization: lowercase + strip everything except a-z and 0-9.
# So "Chg. Cap.(mAh)" → "chgcapmah", matched by pattern "chgcap".
# ---------------------------------------------------------------------------

CYCLER_SIGNATURES: Dict[str, List[List[str]]] = {
    "neware": [
        ["cycleindex", "cycleid", "cycleno"],
        ["stepindex",  "stepid",  "stepno"],
        ["chgcap",  "chargecap",  "capchg",  "capcharge"],
        ["dchgcap", "dischargecap", "capdchg", "capdischarge", "discap"],
        ["chgdchgeff", "efficiency", "coulombiceff", "ceeff"],
    ],
    "biologic": [
        ["ewev", "ewe"],
        ["ima", "imA"],
        ["times", "timesec"],
        ["qcharge", "chargeq"],
        ["oxred"],
    ],
    "arbin": [
        ["cycleindex", "cycleid"],
        ["chargecapacityah", "chgcapah"],
        ["dischargecapacityah", "discapah"],
        ["internalresistance", "dcir"],
    ],
}


# ---------------------------------------------------------------------------
# Map cycler name → reader class.
# ---------------------------------------------------------------------------

def _CYCLER_READER_MAP() -> Dict[str, Optional[Type[BaseReader]]]:
    from cellseer.readers.cycling.neware_reader import NewareReader
    from cellseer.readers.cycling.biologic_reader import BiologicReader
    from cellseer.readers.cycling.arbin_reader import ArbinReader
    return {
        "neware":   NewareReader,
        "biologic": BiologicReader,
        "arbin":    ArbinReader,
    }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def detect_cycler(filepath: Path | str) -> Tuple[Optional[str], float]:
    """
    Detect the cycler manufacturer by scoring column headers against
    CYCLER_SIGNATURES.

    Returns
    -------
    (cycler_name, confidence)
        cycler_name : str or None if no match found
        confidence  : 0.0–1.0  (fraction of signature concepts matched)
    """
    headers = _read_headers(Path(filepath))
    norm_headers = [_normalize(h) for h in headers]

    best_cycler: Optional[str] = None
    best_score = 0.0

    for cycler, concepts in CYCLER_SIGNATURES.items():
        matched = sum(
            any(alt in norm_h for alt in alternatives for norm_h in norm_headers)
            for alternatives in concepts
        )
        score = matched / len(concepts)
        if score > best_score:
            best_score = score
            best_cycler = cycler

    if best_score == 0.0:
        return None, 0.0
    return best_cycler, round(best_score, 2)


def detect_reader(
    filepath: Path | str,
    min_confidence: float = 0.4,
    **reader_kwargs,
) -> BaseReader:
    """
    Detect the cycler and return a ready-to-use reader instance.

    Parameters
    ----------
    filepath : Path | str
        Path to the data file.
    min_confidence : float
        Minimum fraction of signature concepts that must match.
    **reader_kwargs
        Forwarded to the reader constructor (e.g. use_pyprobe=False).
    """
    filepath = Path(filepath)
    cycler, confidence = detect_cycler(filepath)

    if cycler is None or confidence < min_confidence:
        detected = f"{cycler} ({confidence:.0%})" if cycler else "none"
        raise ValueError(
            f"Could not confidently detect data format for {filepath.name}. "
            f"Best match: {detected}. "
            f"Available formats: {list(CYCLER_SIGNATURES.keys())}"
        )

    reader_map = _CYCLER_READER_MAP()
    reader_cls = reader_map.get(cycler)

    if reader_cls is None:
        raise NotImplementedError(
            f"Detected data format '{cycler}' (confidence {confidence:.0%}) "
            f"but no reader is implemented yet."
        )

    return reader_cls(input_paths=filepath, **reader_kwargs)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _normalize(text: str) -> str:
    """Lowercase and strip everything except a-z and 0-9."""
    return re.sub(r'[^a-z0-9]', '', text.lower())


def _read_headers(filepath: Path) -> List[str]:
    """
    Extract column headers from a file without loading data rows.
    Supports .xlsx, .xls, .csv, .parquet.
    """
    suffix = filepath.suffix.lower()

    # BioLogic binary/text formats — detect by extension; header parsing is unreliable
    if suffix in (".mpt", ".mpr"):
        return ["ox_red", "ecell", "capacitycharge", "capacitydischarge", "time"]

    if suffix in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # Try "record" sheet first (Neware), then all sheets
        sheet_names = (
            ["record"] + [s for s in wb.sheetnames if s != "record"]
            if "record" in wb.sheetnames
            else wb.sheetnames
        )
        best_headers: List[str] = []
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            # Scan the first 20 rows; use the row with the most non-empty cells
            for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if len(cells) > len(best_headers):
                    best_headers = cells
            if best_headers:
                break
        wb.close()
        return best_headers

    if suffix == ".csv":
        with open(filepath, encoding="utf-8", errors="ignore") as f:
            first_line = f.readline().strip()
        return [h.strip() for h in first_line.split(",")]

    if suffix == ".parquet":
        import polars as pl
        return pl.read_parquet_schema(filepath).names()

    return []
