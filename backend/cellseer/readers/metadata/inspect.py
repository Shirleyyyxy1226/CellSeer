"""Read and inspect metadata spreadsheets (Excel / CSV)."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional

from cellseer.analysis.metadata.columns import (
    candidate_id_no_headers,
    candidate_key_headers,
    choose_default_id_no_header,
    choose_default_primary_key,
    detect_header_row_index,
)
from cellseer.analysis.metadata.fields import normalize_key

DISPLAY_NAME_TEMPLATE = "{cathode}_{anode}_R{repeat}_ID{id_no}"


def inspect_metadata_file(
    filepath: Path,
    preferred_sheet: Optional[str] = None,
    preferred_primary_key: Optional[str] = None,
    preferred_id_no: Optional[str] = None,
) -> Dict[str, Any]:
    """Inspect metadata file and return rows + key-candidate diagnostics."""
    suffix = filepath.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return _inspect_excel_metadata(
            filepath,
            preferred_sheet=preferred_sheet,
            preferred_primary_key=preferred_primary_key,
            preferred_id_no=preferred_id_no,
        )
    if suffix == ".csv":
        return _inspect_csv_metadata(filepath, preferred_id_no=preferred_id_no)
    raise ValueError(
        f"Unsupported metadata file format: {filepath.suffix!r}. "
        "Expected .xlsx, .xls, or .csv."
    )


def read_metadata_rows(
    filepath: Path | str,
    sheet_name: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Return parsed metadata rows using the same header detection as inspect."""
    parsed = inspect_metadata_file(
        Path(filepath),
        preferred_sheet=sheet_name,
    )
    return parsed.get("rows", [])


def _inspect_excel_metadata(
    filepath: Path,
    preferred_sheet: Optional[str],
    preferred_primary_key: Optional[str],
    preferred_id_no: Optional[str],
) -> Dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        inspect_sheets = (
            [preferred_sheet] if preferred_sheet and preferred_sheet in sheet_names else sheet_names
        )

        best: Optional[Dict[str, Any]] = None
        for sheet_name in inspect_sheets:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx = detect_header_row_index(rows)
            headers = [str(h).strip() if h is not None else "" for h in rows[header_idx]]
            records = [
                {h: v for h, v in zip(headers, row) if h}
                for row in rows[header_idx + 1 :]
                if any(v is not None and str(v).strip() for v in row)
            ]
            candidates = candidate_key_headers(headers)
            id_no_candidates = candidate_id_no_headers(headers, records)
            norm_headers = {normalize_key(h) for h in headers if h}
            sheet_score = len(candidates) * 10 + len(records)
            if "cellname" in norm_headers and "cellno" in norm_headers:
                sheet_score += 100
            if "cellname" in norm_headers:
                sheet_score += 60
            if "cellno" in norm_headers or "idno" in norm_headers:
                sheet_score += 50
            if "cathode" in norm_headers and "anode" in norm_headers:
                sheet_score += 25
            if preferred_primary_key and normalize_key(preferred_primary_key) in norm_headers:
                sheet_score += 120

            payload = {
                "sheet_name": sheet_name,
                "sheet_names": sheet_names,
                "header_row_index": header_idx,
                "headers": headers,
                "rows": records,
                "key_candidates": candidates,
                "default_primary_key": choose_default_primary_key(candidates, headers),
                "requires_primary_key_choice": len(candidates) > 1,
                "id_no_candidates": id_no_candidates,
                "default_id_no_header": choose_default_id_no_header(
                    id_no_candidates, headers, preferred_id_no=preferred_id_no
                ),
                "requires_id_no_choice": len(id_no_candidates) > 1,
                "file_type": "metadata",
                "record_count": len(records),
                "display_name_template": DISPLAY_NAME_TEMPLATE,
            }
            if best is None or sheet_score > best["score"]:
                payload["score"] = sheet_score
                best = payload

        if best is None:
            return {
                "sheet_name": preferred_sheet or wb.active.title,
                "sheet_names": sheet_names,
                "header_row_index": 0,
                "headers": [],
                "rows": [],
                "key_candidates": [],
                "default_primary_key": None,
                "requires_primary_key_choice": False,
                "id_no_candidates": [],
                "default_id_no_header": None,
                "requires_id_no_choice": False,
                "file_type": "metadata",
                "record_count": 0,
                "display_name_template": DISPLAY_NAME_TEMPLATE,
            }

        best.pop("score", None)
        return best
    finally:
        wb.close()


def _inspect_csv_metadata(filepath: Path, preferred_id_no: Optional[str] = None) -> Dict[str, Any]:
    with open(filepath, encoding="utf-8", errors="ignore", newline="") as fh:
        reader = csv.DictReader(fh)
        rows = [dict(row) for row in reader]
        headers = list(reader.fieldnames or [])
    candidates = candidate_key_headers(headers)
    id_no_candidates = candidate_id_no_headers(headers, rows)
    return {
        "sheet_name": None,
        "sheet_names": [],
        "header_row_index": 0,
        "headers": headers,
        "rows": rows,
        "key_candidates": candidates,
        "default_primary_key": choose_default_primary_key(candidates, headers),
        "requires_primary_key_choice": len(candidates) > 1,
        "id_no_candidates": id_no_candidates,
        "default_id_no_header": choose_default_id_no_header(
            id_no_candidates, headers, preferred_id_no=preferred_id_no
        ),
        "requires_id_no_choice": len(id_no_candidates) > 1,
        "file_type": "metadata",
        "record_count": len(rows),
        "display_name_template": DISPLAY_NAME_TEMPLATE,
    }
