"""
cellseer/ingest.py — Top-level helpers for loading data into a CellSeer database.

Two entry points
----------------
ingest_metadata(filepath, db)
    Read a metadata spreadsheet (Excel or CSV) and upsert every row as a Cell
    (metadata only, no test data yet).

ingest_cycling_file(filepath, db, dataset_name="cycling", **reader_kwargs)
    Detect the cycler format from file content, identify which cell the file
    belongs to via the numeric ``id_no`` prefix in the filename, read the data,
    and store it in the DB.  Accepts one file or a list of files (continuation
    tests are concatenated automatically).

Typical workflow
----------------
>>> from cellseer import SQLiteBackend
>>> from cellseer.ingest import ingest_metadata, ingest_cycling_file
>>>
>>> db = SQLiteBackend("cellseer.db")
>>>
>>> # 1. Load cell metadata
>>> ingest_metadata("data/metadata/cells.xlsx", db)
>>>
>>> # 2. Single file
>>> ingest_cycling_file("data/raw/1073_Rate testing.xlsx", db)
>>>
>>> # 3. Two continuation files for the same cell
>>> ingest_cycling_file(["data/raw/1073_Part1.xlsx", "data/raw/1073_Part2.xlsx"], db)
>>>
>>> # 4. Batch-load a whole folder
>>> from pathlib import Path
>>> for f in Path("data/raw").glob("*.xlsx"):
...     ingest_cycling_file(f, db)
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import TYPE_CHECKING, Any, Callable, Dict, List, Optional, Union

if TYPE_CHECKING:
    from cellseer.db.base import DBBackend


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def ingest_metadata(
    filepath: Path | str,
    db: "DBBackend",
    primary_key_header: Optional[str] = None,
    id_no_header: Optional[str] = None,
    display_name_template: str = "{cathode}_{anode}_R{repeat}_ID{id_no}",
    metadata_sheet: Optional[str] = None,
    progress_cb: Optional[Callable[[int, str], None]] = None,
    project_id: Optional[str] = None,
) -> List[str]:
    """
    Read a metadata file and upsert each row as a Cell (metadata only).

    Parameters
    ----------
    filepath : Path | str
        Path to an Excel (.xlsx, .xls) or CSV file.
    db : DBBackend
        Open database backend (SQLiteBackend or DuckDBBackend).

    Returns
    -------
    List[str]
        cell_ids that were upserted.
    """
    from cellseer.cell import Cell
    from cellseer.metadata import CellMetadata

    parsed = inspect_metadata_file(
        Path(filepath),
        preferred_sheet=metadata_sheet,
        preferred_primary_key=primary_key_header,
        preferred_id_no=id_no_header,
    )
    rows = parsed["rows"]
    id_no_candidates = parsed.get("id_no_candidates", []) or []
    selected_id_no_header = id_no_header or parsed.get("default_id_no_header")
    if not id_no_candidates:
        raise ValueError(
            "No numeric link-key column found. Please add/select an ID column for data linkage."
        )
    if len(id_no_candidates) > 1 and not id_no_header:
        raise ValueError(
            "Multiple numeric link-key columns found. Please select an ID column for data linkage."
        )
    if selected_id_no_header and selected_id_no_header not in parsed.get("headers", []):
        raise ValueError(
            f"Selected link-key column '{selected_id_no_header}' is not present in metadata headers."
        )
    upserted: List[str] = []

    skipped_rows = 0
    total_rows = len(rows)
    progress_step = max(1, total_rows // 20) if total_rows else 1
    if progress_cb:
        progress_cb(45, f"Upserting cells... 0/{total_rows}")
    for idx, row in enumerate(rows, start=1):
        try:
            meta = CellMetadata.from_dict(
                row,
                primary_key_header=primary_key_header,
                id_no_header=selected_id_no_header,
                display_name_template=display_name_template,
            )
        except Exception as exc:
            # Metadata sheets may contain decorative/header/footer rows; skip invalid rows.
            skipped_rows += 1
            continue

        cell = Cell(metadata=meta)
        if _id_no_conflicts(db, meta.id_no, meta.cell_id, project_id=project_id):
            cell.metadata.id_no = None
        db.upsert_cell(cell)
        upserted.append(meta.cell_id)

        if progress_cb and (idx == 1 or idx % progress_step == 0 or idx == total_rows):
            # Keep this in 45..95 so caller can reserve final 100 for completion.
            pct = 45 + int((idx / max(total_rows, 1)) * 50)
            progress_cb(
                min(pct, 95),
                f"Upserting cells... {idx}/{total_rows} (imported {len(upserted)})",
            )

    if not upserted:
        raise ValueError(
            "No valid metadata rows found. Ensure a valid primary-key column "
            "(for example: Cell name, ID, or Cell_ID) is selected."
        )

    if skipped_rows:
        print(f"Skipped {skipped_rows} invalid metadata rows during import.")
    if progress_cb:
        progress_cb(95, f"Upserting cells... {total_rows}/{total_rows} (imported {len(upserted)})")

    return upserted


def ingest_cycling_file(
    filepath: Union[Path, str, List[Union[Path, str]]],
    db: "DBBackend",
    dataset_name: str = "cycling",
    min_confidence: float = 0.4,
    project_id: Optional[str] = None,
    **reader_kwargs,
) -> str:
    """
    Detect the cycler format, find the matching cell in the DB, read the data,
    and store it.

    Accepts one file or a list of files for the same cell (continuation tests).
    All files must share the same id_no prefix — they are concatenated in order.

    Parameters
    ----------
    filepath : Path | str | list
        One file or a list of files (.xlsx, .csv, .parquet).
    db : DBBackend
        Open database backend.
    dataset_name : str
        Key under which the data is stored (default ``"cycling"``).
    min_confidence : float
        Minimum signature-match confidence required by detect_reader().
    **reader_kwargs
        Forwarded to the reader (e.g. ``use_pyprobe=False``).

    Returns
    -------
    str
        The ``cell_id`` of the cell that was updated.
    """
    from cellseer.cell import Cell
    from cellseer.readers.cycling.detect import detect_reader

    # Normalise to a list of Paths
    if isinstance(filepath, (str, Path)):
        filepaths = [Path(filepath)]
    else:
        filepaths = [Path(f) for f in filepath]

    primary = filepaths[0]

    # 1. Extract id_no from the first filename
    id_no = _id_no_from_filename(primary)

    # 2. Find matching cell in DB (returns None if not found)
    cell_id = _find_cell_id(id_no, db, project_id=project_id)

    # 3. Load or create cell
    if cell_id is None:
        from cellseer.metadata import CellMetadata
        cell_id = str(id_no)
        print(
            f"  Warning: no metadata found for id_no={id_no}. "
            f"Creating placeholder cell '{cell_id}'. "
            "Run ingest_metadata() to add full metadata later."
        )
        cell = Cell(metadata=CellMetadata(cell_id=cell_id, id_no=id_no))
    else:
        try:
            cell = Cell.load(db, cell_id)
        except KeyError:
            from cellseer.metadata import CellMetadata
            cell = Cell(metadata=CellMetadata(cell_id=cell_id, id_no=id_no))

    # 4. Detect cycler (for logging), then get reader
    from cellseer.readers.cycling.detect import detect_cycler
    cycler, confidence = detect_cycler(primary)

    reader = detect_reader(
        primary,
        min_confidence=min_confidence,
        info={"cell_id": cell_id},
        **reader_kwargs,
    )
    reader.input_paths = filepaths   # reader concatenates them in order
    raw = reader.read()
    raw.info.setdefault("cell_id", cell_id)
    cell.datasets[dataset_name] = raw

    # 5. Try computing dQ/dV and dV/dQ for both discharge and charge.
    # Do not reject valid cycling uploads if differentiation is unavailable.
    for direction in ("discharge", "charge"):
        try:
            cell.compute_dqdv(
                cycling_dataset=dataset_name,
                direction=direction,
                output_prefix=f"{direction}_",
            )
        except ValueError as exc:
            print(f"  Warning: ICA/DVQ ({direction}) skipped for cell '{cell_id}': {exc}")

    # 6. Persist everything in one write
    cell.save(db)

    # 7. Audit log (best-effort — backends without log_ingest() are a no-op)
    try:
        db.log_ingest(cell_id, filepaths, cycler=cycler, confidence=confidence)
    except Exception:
        pass

    return cell_id


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def inspect_metadata_file(
    filepath: Path,
    preferred_sheet: Optional[str] = None,
    preferred_primary_key: Optional[str] = None,
    preferred_id_no: Optional[str] = None,
) -> Dict[str, Any]:
    """Inspect metadata file and return rows + key-candidate diagnostics."""
    suffix = filepath.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return _inspect_excel_metadata(
            filepath,
            preferred_sheet=preferred_sheet,
            preferred_primary_key=preferred_primary_key,
            preferred_id_no=preferred_id_no,
        )
    if suffix == ".csv":
        return _inspect_csv_metadata(filepath, preferred_id_no=preferred_id_no)
    raise ValueError(
        f"Unsupported metadata file format: {filepath.suffix!r}. "
        "Expected .xlsx, .xls, or .csv."
    )


def _norm_header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


_KEY_HEADER_ALIASES = {
    "cellname",
    "cellid",
    "id",
    "sampleid",
    "barcode",
    "name",
}

_ID_NO_HEADER_ALIASES = {
    "idno",
    "idnumber",
    "cellno",
    "celno",
    "number",
}

_HEADER_HINTS = {
    "id",
    "name",
    "cellname",
    "cellid",
    "idno",
    "cellno",
    "cathode",
    "anode",
    "separatortype",
    "electrolyte",
    "repeat",
    "batch",
}


def _candidate_key_headers(headers: List[str]) -> List[str]:
    out: List[str] = []
    seen: set[str] = set()
    for h in headers:
        nh = _norm_header_key(h)
        if nh in _KEY_HEADER_ALIASES and h not in seen:
            out.append(h)
            seen.add(h)
    return out


def _choose_default_primary_key(candidates: List[str], headers: List[str]) -> Optional[str]:
    if not headers:
        return None
    by_norm = {_norm_header_key(h): h for h in headers if h}
    for preferred in ("cellname", "id", "cellid", "sampleid", "barcode"):
        if preferred in by_norm:
            return by_norm[preferred]
    return candidates[0] if candidates else None


def _coerce_int(value: object) -> Optional[int]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        if re.fullmatch(r"[+-]?\d+", s):
            return int(s)
        f = float(s)
        if f.is_integer():
            return int(f)
    except Exception:
        return None
    return None


def _looks_like_id_no_column(norm_header: str) -> bool:
    if norm_header in _ID_NO_HEADER_ALIASES:
        return True
    return ("id" in norm_header) or ("no" in norm_header) or ("number" in norm_header)


def _candidate_id_no_headers(headers: List[str], rows: List[Dict[str, Any]]) -> List[str]:
    candidates: List[str] = []
    for h in headers:
        if not h:
            continue
        nh = _norm_header_key(h)
        if not _looks_like_id_no_column(nh):
            continue
        if nh in {"batch", "repeat", "cycle", "step"}:
            continue
        values = []
        for r in rows:
            v = r.get(h)
            if v is None or str(v).strip() == "":
                continue
            values.append(v)
            if len(values) >= 20:
                break
        if not values:
            continue
        parsed = [_coerce_int(v) for v in values]
        valid = [x for x in parsed if x is not None]
        if len(valid) / len(values) >= 0.95 and sum(1 for x in valid if x >= 1) / max(len(valid), 1) >= 0.9:
            candidates.append(h)
    return candidates


def _choose_default_id_no_header(
    candidates: List[str],
    headers: List[str],
    preferred_id_no: Optional[str] = None,
) -> Optional[str]:
    if preferred_id_no:
        for h in headers:
            if _norm_header_key(h) == _norm_header_key(preferred_id_no):
                return h
    if not candidates:
        return None
    by_norm = {_norm_header_key(h): h for h in candidates}
    for preferred in ("idno", "idnumber", "number", "celno", "cellno"):
        if preferred in by_norm:
            return by_norm[preferred]
    return candidates[0]


def _detect_header_idx(rows: List[tuple], scan_limit: int = 60) -> int:
    best_idx = 0
    best_score = -1
    for i, candidate in enumerate(rows[:scan_limit]):
        labels = [
            _norm_header_key(v) for v in candidate if v is not None and str(v).strip()
        ]
        if not labels:
            continue
        score = sum(1 for x in labels if x in _HEADER_HINTS) + len(set(labels)) * 0.1
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def _inspect_excel_metadata(
    filepath: Path,
    preferred_sheet: Optional[str],
    preferred_primary_key: Optional[str],
    preferred_id_no: Optional[str],
) -> Dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        inspect_sheets = (
            [preferred_sheet] if preferred_sheet and preferred_sheet in sheet_names else sheet_names
        )

        best: Optional[Dict[str, Any]] = None
        for sheet_name in inspect_sheets:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx = _detect_header_idx(rows)
            headers = [str(h).strip() if h is not None else "" for h in rows[header_idx]]
            records = [
                {h: v for h, v in zip(headers, row) if h}
                for row in rows[header_idx + 1 :]
                if any(v is not None and str(v).strip() for v in row)
            ]
            candidates = _candidate_key_headers(headers)
            id_no_candidates = _candidate_id_no_headers(headers, records)
            norm_headers = {_norm_header_key(h) for h in headers if h}
            sheet_score = len(candidates) * 10 + len(records)
            if "cellname" in norm_headers and "cellno" in norm_headers:
                # Strong indicator of assembly metadata tables (e.g. CoinCellAssemble).
                sheet_score += 100
            if "cellname" in norm_headers:
                sheet_score += 60
            if "cellno" in norm_headers or "idno" in norm_headers:
                sheet_score += 50
            if "cathode" in norm_headers and "anode" in norm_headers:
                sheet_score += 25
            if preferred_primary_key and _norm_header_key(preferred_primary_key) in norm_headers:
                sheet_score += 120

            payload = {
                "sheet_name": sheet_name,
                "sheet_names": sheet_names,
                "header_row_index": header_idx,
                "headers": headers,
                "rows": records,
                "key_candidates": candidates,
                "default_primary_key": _choose_default_primary_key(candidates, headers),
                "requires_primary_key_choice": len(candidates) > 1,
                "id_no_candidates": id_no_candidates,
                "default_id_no_header": _choose_default_id_no_header(
                    id_no_candidates, headers, preferred_id_no=preferred_id_no
                ),
                "requires_id_no_choice": len(id_no_candidates) > 1,
                "file_type": "metadata",
                "record_count": len(records),
                "display_name_template": "{cathode}_{anode}_R{repeat}_ID{id_no}",
            }
            if best is None or sheet_score > best["score"]:
                payload["score"] = sheet_score
                best = payload

        if best is None:
            return {
                "sheet_name": preferred_sheet or wb.active.title,
                "sheet_names": sheet_names,
                "header_row_index": 0,
                "headers": [],
                "rows": [],
                "key_candidates": [],
                "default_primary_key": None,
                "requires_primary_key_choice": False,
                "id_no_candidates": [],
                "default_id_no_header": None,
                "requires_id_no_choice": False,
                "file_type": "metadata",
                "record_count": 0,
                "display_name_template": "{cathode}_{anode}_R{repeat}_ID{id_no}",
            }

        best.pop("score", None)
        return best
    finally:
        wb.close()


def _inspect_csv_metadata(filepath: Path, preferred_id_no: Optional[str] = None) -> Dict[str, Any]:
    with open(filepath, encoding="utf-8", errors="ignore", newline="") as fh:
        reader = csv.DictReader(fh)
        rows = [dict(row) for row in reader]
        headers = list(reader.fieldnames or [])
    candidates = _candidate_key_headers(headers)
    id_no_candidates = _candidate_id_no_headers(headers, rows)
    return {
        "sheet_name": None,
        "sheet_names": [],
        "header_row_index": 0,
        "headers": headers,
        "rows": rows,
        "key_candidates": candidates,
        "default_primary_key": _choose_default_primary_key(candidates, headers),
        "requires_primary_key_choice": len(candidates) > 1,
        "id_no_candidates": id_no_candidates,
        "default_id_no_header": _choose_default_id_no_header(
            id_no_candidates, headers, preferred_id_no=preferred_id_no
        ),
        "requires_id_no_choice": len(id_no_candidates) > 1,
        "file_type": "metadata",
        "record_count": len(rows),
        "display_name_template": "{cathode}_{anode}_R{repeat}_ID{id_no}",
    }


def _id_no_from_filename(filepath: Path) -> int:
    """
    Extract id_no from filename with flexible patterns.

    Supported examples:
    - '1073_Rate testing.xlsx' -> 1073
    - 'CEL-100-NMC622-Graphite.xlsx' -> 100
    - 'P025-CEL-9.xlsx' -> 9
    """
    stem = filepath.stem

    # 1) Classic numeric prefix.
    m = re.match(r"^(\d+)", stem)
    if m:
        return int(m.group(1))

    # 2) Common cell tokens: CEL-100, CELL_100, P025-CEL-9, etc.
    m = re.search(r"(?:^|[^a-z0-9])(cel|cell)[-_ ]*(\d+)(?:[^a-z0-9]|$)", stem, flags=re.IGNORECASE)
    if m:
        return int(m.group(2))

    # 3) Generic fallback: first integer group in the filename.
    m = re.search(r"(\d+)", stem)
    if m:
        return int(m.group(1))

    raise ValueError(
        f"Cannot extract id_no from filename {filepath.name!r}. "
        "Supported patterns include numeric prefix (e.g. '1073_...') or cell token "
        "(e.g. 'CEL-100-...')."
    )


def _find_cell_id(id_no: int, db: "DBBackend", project_id: Optional[str] = None) -> Optional[str]:
    """Return the cell_id whose id_no matches, or None if not found."""
    if hasattr(db, "_connect"):
        pid = project_id or getattr(db, "project_id", "default")
        conn = db._connect()
        row = conn.execute(
            "SELECT cell_id FROM cell WHERE project_id = ? AND id_no = ?", (pid, id_no)
        ).fetchone()
        conn.close()
        if row is None:
            return None
        return row["cell_id"] if hasattr(row, "keys") else row[0]

    from cellseer.cell import Cell
    for cell_id in db.list_cells():
        try:
            cell = Cell.load(db, cell_id)
            if cell.metadata.id_no == id_no:
                return cell_id
        except Exception:
            continue
    return None


def _id_no_conflicts(
    db: "DBBackend",
    id_no: Optional[int],
    cell_id: str,
    project_id: Optional[str] = None,
) -> bool:
    if id_no is None:
        return False
    if not hasattr(db, "_connect"):
        return False
    pid = project_id or getattr(db, "project_id", "default")
    conn = db._connect()
    try:
        row = conn.execute(
            "SELECT cell_id FROM cell WHERE project_id = ? AND id_no = ?",
            (pid, id_no),
        ).fetchone()
        if row is None:
            return False
        existing_cell_id = row["cell_id"] if hasattr(row, "keys") else row[0]
        return existing_cell_id != cell_id
    finally:
        conn.close()
